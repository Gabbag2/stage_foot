from PIL import Image, ImageEnhance 
from utils import couper, mise_dans_excel, ameliorer_image, creer_toutes_les_cases
import itertools
from tqdm import tqdm
import pytesseract
import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Chemin vers votre image
image_path = r"C:\Users\gabri\Desktop\stage_CNRS\image.png"

image = ameliorer_image(image_path) # ouvre l'image + augmente le contraste

dictionnaire_de_renomage = {'mot':'mot2'} # a faire pour toutes les erreurs possibles !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

pixel_case = creer_toutes_les_cases() # creer un tableau avec les pixels de chaque case

lettre = ('A','B','C','D','E','F','G','H','I')
# Charger le fichier Excel existant ou en créer un nouveau
output_path = r"C:\Users\gabri\Desktop\stage_CNRS\output2.xlsx" #mettre nouvel output§§§§§§§§§§§§§§§§§§§§§§§§§§§§§§§
try:
    workbook = load_workbook(output_path)
except FileNotFoundError:
    workbook = Workbook()

sheet = workbook.active


for i in tqdm(range(len(pixel_case))):
    left = pixel_case[i][0][1]
    upper = pixel_case[i][0][0]
    right = pixel_case[i][1][1]
    lower = pixel_case[i][1][0]
    
    chiffre_excel = lettre[i%9]
    case_excel = i//9+1
    image_cropped = couper(image,left, upper, right, lower)
    mise_dans_excel(image_cropped,chiffre_excel,case_excel,dictionnaire_de_renomage,sheet)
    """if i < 5 :
        image_cropped.show()
    if chiffre_excel == "F":
        if case_excel == 21 or case_excel == 20:
            image_cropped.show()"""
# Sauvegarder le fichier Excel
workbook.save(output_path)

